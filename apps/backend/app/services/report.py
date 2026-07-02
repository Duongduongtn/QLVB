"""Báo cáo — G2 xuất sổ CV đi/đến theo NĐ 30/2020 (Excel) + thống kê dashboard (G1-lite).

`build_register_xlsx` sinh file Excel đúng cột Phụ lục III NĐ 30/2020 (header tiếng Việt có
dấu), openpyxl import TRỄ. `dashboard_stats` trả KPI + 12 tháng CV đi/đến (đếm thật).
"""

from __future__ import annotations

from datetime import UTC, date, datetime, time, timedelta
from typing import Any

from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session

from app.core.errors import ValidationFailed
from app.models.document_type import DocumentType
from app.models.incoming_document import IncomingDocument
from app.models.organization import Organization
from app.models.outgoing_document import OutgoingDocument, OutgoingRecipient
from app.models.processing_task import ProcessingTask
from app.models.signature import Signature
from app.models.signing_profile import SigningProfile
from app.models.unit import Unit

BOOKS = ("di_gdnn", "di_dvdl", "den")

# Cột NĐ 30/2020 — Phụ lục III (sổ đăng ký văn bản đi / đến).
_COLS_DI = [
    ("stt", "Số TT", 8),
    ("number", "Số, ký hiệu văn bản", 22),
    ("issue_date", "Ngày tháng văn bản", 16),
    ("subject", "Tên loại và trích yếu nội dung văn bản", 50),
    ("signer", "Người ký", 22),
    ("recipients", "Nơi nhận văn bản", 32),
    ("copies", "Số lượng bản", 12),
    ("note", "Ghi chú", 20),
]
_COLS_DEN = [
    ("stt", "Số đến", 8),
    ("arrival_date", "Ngày đến", 14),
    ("reference", "Số, ký hiệu", 20),
    ("doc_date", "Ngày tháng văn bản", 16),
    ("sender", "Cơ quan, tổ chức ban hành", 30),
    ("subject", "Tên loại và trích yếu nội dung", 46),
    ("assignee", "Đơn vị/người nhận xử lý", 24),
    ("note", "Ghi chú", 20),
]

_BOOK_TITLE = {
    "di_gdnn": "SỔ ĐĂNG KÝ VĂN BẢN ĐI — GDNN",
    "di_dvdl": "SỔ ĐĂNG KÝ VĂN BẢN ĐI — DVDL",
    "den": "SỔ ĐĂNG KÝ VĂN BẢN ĐẾN (chung 2 đơn vị)",
}
_STATUS_NOTE = {"cancelled": "Đã huỷ"}


def _fmt_d(d: date | None) -> str:
    return d.strftime("%d/%m/%Y") if d else ""


def _excel_safe(v: Any) -> Any:
    """Chống Excel/CSV formula injection (TDD §10.2): field tự do bắt đầu bằng = + - @ TAB CR
    → prepend `'` để Excel coi là TEXT, không diễn giải thành công thức (trích yếu/tên cơ quan
    do bên ngoài soạn / OCR có thể chứa)."""
    if isinstance(v, str) and v[:1] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + v
    return v


def _year_range(year: int) -> tuple[date, date]:
    return date(year, 1, 1), date(year + 1, 1, 1)


def _outgoing_rows(db: Session, unit_code: str, year: int) -> list[dict[str, Any]]:
    unit = db.scalar(select(Unit).where(Unit.code == unit_code))
    if unit is None:
        return []
    lo, hi = _year_range(year)
    docs = db.scalars(
        select(OutgoingDocument)
        .where(
            OutgoingDocument.unit_id == unit.id,
            OutgoingDocument.deleted_at.is_(None),
            OutgoingDocument.number_int.is_not(None),
            OutgoingDocument.issue_date >= lo,
            OutgoingDocument.issue_date < hi,
        )
        .order_by(OutgoingDocument.number_int.asc())
    ).all()
    if not docs:
        return []
    # Batch tránh N+1: nạp người ký + nơi nhận 1 lần.
    profile_ids = {d.signing_profile_id for d in docs if d.signing_profile_id is not None}
    signer_map: dict[int, str] = {}
    if profile_ids:
        signer_map = {
            pid: name
            for pid, name in db.execute(
                select(SigningProfile.id, Signature.full_name)
                .join(Signature, SigningProfile.signature_id == Signature.id)
                .where(SigningProfile.id.in_(profile_ids))
            ).all()
        }
    recip_map: dict[int, list[str]] = {}
    for oid, name in db.execute(
        select(OutgoingRecipient.outgoing_id, Organization.full_name)
        .join(Organization, OutgoingRecipient.organization_id == Organization.id)
        .where(OutgoingRecipient.outgoing_id.in_([d.id for d in docs]))
    ).all():
        recip_map.setdefault(oid, []).append(name)

    rows: list[dict[str, Any]] = []
    for i, d in enumerate(docs, start=1):
        rows.append({
            "stt": i,
            "number": d.number or "",
            "issue_date": _fmt_d(d.issue_date),
            "subject": d.subject or "",
            "signer": signer_map.get(d.signing_profile_id, "") if d.signing_profile_id else "",
            "recipients": "; ".join(recip_map.get(d.id, [])),
            "copies": "",
            "note": _STATUS_NOTE.get(d.status, ""),
        })
    return rows


def _incoming_rows(db: Session, year: int) -> list[dict[str, Any]]:
    lo, hi = _year_range(year)
    docs = db.scalars(
        select(IncomingDocument)
        .where(
            IncomingDocument.deleted_at.is_(None),
            IncomingDocument.number_int.is_not(None),
            IncomingDocument.created_at >= lo,
            IncomingDocument.created_at < hi,
        )
        .order_by(IncomingDocument.number_int.asc())
    ).all()
    if not docs:
        return []
    sender_ids = {d.sender_org_id for d in docs if d.sender_org_id is not None}
    sender_map: dict[int, str] = {}
    if sender_ids:
        sender_map = {
            oid: name
            for oid, name in db.execute(
                select(Organization.id, Organization.full_name).where(Organization.id.in_(sender_ids))
            ).all()
        }
    rows: list[dict[str, Any]] = []
    for d in docs:
        rows.append({
            # "Số đến" = số đã cấp THẬT (number), KHÔNG phải thứ tự dòng (số có thể nhảy).
            "stt": d.number or "",
            "arrival_date": _fmt_d(d.created_at.date() if d.created_at else None),
            "reference": d.reference_number or "",
            "doc_date": _fmt_d(d.document_date),
            "sender": sender_map.get(d.sender_org_id, "") if d.sender_org_id else "",
            "subject": d.subject or "",
            "assignee": "",
            "note": _STATUS_NOTE.get(d.status, ""),
        })
    return rows


_SHEET_NAME = {"di_gdnn": "So di GDNN", "di_dvdl": "So di DVDL", "den": "So den"}


def _fill_register_sheet(ws: Any, *, book: str, rows: list[dict[str, Any]], year: int) -> None:
    """Đổ 1 sheet sổ đăng ký NĐ 30/2020 (tiêu đề + header + dữ liệu + freeze) vào `ws`.

    Tách riêng để dùng chung cho file 1 sổ (G2) lẫn `index.xlsx` 3 sổ (G4 export ZIP)."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    cols = _COLS_DEN if book == "den" else _COLS_DI
    ncol = len(cols)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    title_cell = ws.cell(row=1, column=1, value=f"{_BOOK_TITLE[book]} — NĂM {year}")
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="FCE8B2")
    header_row = 3
    for c, (_key, label, width) in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=c, value=label)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[header_row].height = 32

    for r, row in enumerate(rows, start=header_row + 1):
        for c, (key, _label, _w) in enumerate(cols, start=1):
            cell = ws.cell(row=r, column=c, value=_excel_safe(row.get(key, "")))
            cell.border = border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=key in ("subject", "recipients", "sender"),
                horizontal="center" if key in ("stt",) else "left",
            )

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _register_rows(db: Session, book: str, year: int) -> list[dict[str, Any]]:
    if book == "den":
        return _incoming_rows(db, year)
    return _outgoing_rows(db, "GDNN" if book == "di_gdnn" else "DVDL", year)


def build_register_xlsx(db: Session, *, year: int, book: str) -> bytes:
    """Sinh Excel sổ đăng ký NĐ 30/2020 (1 sổ). Trả bytes .xlsx."""
    if book not in BOOKS:
        raise ValidationFailed("Loại sổ không hợp lệ")

    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None  # Workbook() mới luôn có sheet active
    ws.title = "So dang ky"
    _fill_register_sheet(ws, book=book, rows=_register_rows(db, book, year), year=year)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_register_workbook_bytes(db: Session, *, year: int) -> bytes:
    """G4 — `index.xlsx` gộp 3 sổ NĐ 30 (đi GDNN / đi DVDL / đến) thành 3 sheet."""
    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    for i, book in enumerate(BOOKS):
        ws = wb.active if i == 0 else wb.create_sheet()
        assert ws is not None
        ws.title = _SHEET_NAME[book]
        _fill_register_sheet(ws, book=book, rows=_register_rows(db, book, year), year=year)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# D6 / E5 — Xuất Excel DANH SÁCH (theo bộ lọc đang xem ở sổ CV đi / đến).
# Khác G2 (sổ NĐ 30 cả năm) + G3 (báo cáo pivot): đây là dump đúng danh sách user
# đang lọc (search/đơn vị/trạng thái…), tái dùng list service để khớp 100% bộ lọc.
# --------------------------------------------------------------------------- #

_LIST_EXPORT_CAP = 5000  # trần dòng chống dựng workbook quá lớn (sổ năm ~800 CV)

_URGENCY_VN = {
    "normal": "Thường", "urgent": "Khẩn", "very_urgent": "Thượng khẩn",
    "express": "Hoả tốc", "express_timed": "Hoả tốc hẹn giờ",
}

_COLS_OUT_LIST = [
    ("stt", "STT", 6),
    ("number", "Số, ký hiệu", 18),
    ("issue_date", "Ngày phát hành", 15),
    ("unit", "Đơn vị", 14),
    ("subject", "Trích yếu", 50),
    ("status", "Trạng thái", 14),
]
_COLS_IN_LIST = [
    ("stt", "STT", 6),
    ("number", "Số đến", 10),
    ("arrival_date", "Ngày đến", 13),
    ("reference", "Số, ký hiệu", 16),
    ("doc_date", "Ngày văn bản", 14),
    ("sender", "Cơ quan gửi", 28),
    ("subject", "Trích yếu", 44),
    ("urgency", "Độ khẩn", 12),
    ("status", "Trạng thái", 13),
]


def _render_list_xlsx(title: str, cols: list[Any], rows: list[dict[str, Any]]) -> bytes:
    """Dựng 1 sheet danh sách (tiêu đề + header + dữ liệu, `_excel_safe` chống injection)."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Danh sach"
    ncol = len(cols)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    tc = ws.cell(row=1, column=1, value=title)
    tc.font = Font(bold=True, size=13)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor="FCE8B2")
    hr = 3
    for c, (_key, label, width) in enumerate(cols, start=1):
        cell = ws.cell(row=hr, column=c, value=label)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[hr].height = 28

    for r, row in enumerate(rows, start=hr + 1):
        for c, (key, _label, _w) in enumerate(cols, start=1):
            cell = ws.cell(row=r, column=c, value=_excel_safe(row.get(key, "")))
            cell.border = border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=key in ("subject", "sender"),
                horizontal="center" if key == "stt" else "left",
            )
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_outgoing_list_xlsx(
    db: Session, *, unit_id: int | None = None, status: str | None = None, q: str | None = None
) -> bytes:
    """D6 — xuất danh sách CV đi đang lọc ra Excel."""
    from app.services.outgoing import list_outgoing

    docs, _ = list_outgoing(db, unit_id=unit_id, status=status, q=q, page=1, size=_LIST_EXPORT_CAP)
    units = {u.id: (u.short_name or u.code) for u in db.scalars(select(Unit)).all()}
    rows = [
        {
            "stt": i,
            "number": d.number or "(chưa cấp số)",
            "issue_date": _fmt_d(d.issue_date),
            "unit": units.get(d.unit_id, ""),
            "subject": d.subject or "",
            "status": _STATUS_VN.get(d.status, d.status),
        }
        for i, d in enumerate(docs, start=1)
    ]
    return _render_list_xlsx("DANH SÁCH CÔNG VĂN ĐI", _COLS_OUT_LIST, rows)


def build_incoming_list_xlsx(
    db: Session,
    *,
    include_manager_only: bool,
    status: str | None = None,
    sender_org_id: int | None = None,
    urgency: str | None = None,
    confidentiality: str | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    q: str | None = None,
) -> bytes:
    """E5 — xuất danh sách CV đến đang lọc ra Excel. Tôn trọng manager_only (NV không thấy)."""
    from app.services.incoming import list_incoming

    docs, _ = list_incoming(
        db,
        include_manager_only=include_manager_only,
        status=status,
        sender_org_id=sender_org_id,
        urgency=urgency,
        confidentiality=confidentiality,
        date_from=date_from,
        date_to=date_to,
        q=q,
        page=1,
        size=_LIST_EXPORT_CAP,
    )
    sender_ids = {d.sender_org_id for d in docs if d.sender_org_id is not None}
    sender_map: dict[int, str] = {}
    if sender_ids:
        sender_map = {
            oid: name
            for oid, name in db.execute(
                select(Organization.id, Organization.full_name).where(Organization.id.in_(sender_ids))
            ).all()
        }
    rows = [
        {
            "stt": i,
            "number": d.number or "",
            "arrival_date": _fmt_d(_vn_date(d.created_at)),
            "reference": d.reference_number or "",
            "doc_date": _fmt_d(d.document_date),
            "sender": sender_map.get(d.sender_org_id, "") if d.sender_org_id else "",
            "subject": d.subject or "",
            "urgency": _URGENCY_VN.get(d.urgency or "", d.urgency or ""),
            "status": _STATUS_VN.get(d.status, d.status),
        }
        for i, d in enumerate(docs, start=1)
    ]
    return _render_list_xlsx("DANH SÁCH CÔNG VĂN ĐẾN", _COLS_IN_LIST, rows)


def dashboard_stats(
    db: Session, *, year: int, today: date, unit_id: int | None = None
) -> dict[str, Any]:
    """Dashboard G1 (manager-only → đếm mọi CV). KPI + 12 tháng CV đi/đến + việc chưa xử lý/quá
    hạn + top cơ quan gửi + cơ cấu loại VB đi. `unit_id` (toggle đơn vị) CHỈ lọc CV ĐI; CV ĐẾN
    + việc xử lý dùng chung 2 đơn vị nên KHÔNG lọc (số việc chưa xử lý/quá hạn giống nhau ở mọi
    view; top cơ quan gửi luôn gồm cả 2)."""
    lo, hi = _year_range(year)

    def _monthly(date_col: Any, conds: list[Any]) -> dict[int, int]:
        month = func.extract("month", date_col)
        rows = db.execute(
            select(month.label("m"), func.count().label("c"))
            .where(*conds, date_col >= lo, date_col < hi)
            .group_by(month)
        ).all()
        return {int(r.m): int(r.c) for r in rows}

    di_conds: list[Any] = [
        OutgoingDocument.deleted_at.is_(None),
        OutgoingDocument.number_int.is_not(None),
    ]
    if unit_id is not None:
        di_conds.append(OutgoingDocument.unit_id == unit_id)
    # CV đến lưu created_at UTC → quy về giờ VN trước khi cắt năm/tháng (chống lệch biên ngày/
    # tháng/năm sát giao thừa, nhất quán G3). CV đi dùng issue_date (Date) nên không cần.
    inc_vn = func.timezone("Asia/Ho_Chi_Minh", IncomingDocument.created_at)
    den_conds = [IncomingDocument.deleted_at.is_(None), IncomingDocument.number_int.is_not(None)]
    di_by_month = _monthly(OutgoingDocument.issue_date, di_conds)
    den_by_month = _monthly(inc_vn, den_conds)

    months = [
        {"month": m, "di": di_by_month.get(m, 0), "den": den_by_month.get(m, 0)}
        for m in range(1, 13)
    ]
    cur_m = today.month if today.year == year else 0

    # ── Việc xử lý chưa xong / quá hạn (E2/E3) — đếm theo CV ĐẾN đã vào sổ (registered, loại
    #    nháp + huỷ), năm tiếp nhận theo giờ VN. Sổ đến dùng CHUNG 2 đơn vị và task KHÔNG còn
    #    gắn đơn vị (phân công 1 người) → số liệu tính chung, không lọc theo đơn vị: "Chưa xử
    #    lý" = còn việc mở HOẶC chưa giao việc nào; "Quá hạn" = có việc mở đã quá deadline.
    reg_conds = [
        IncomingDocument.deleted_at.is_(None),
        IncomingDocument.status == "registered",
        inc_vn >= lo,
        inc_vn < hi,
    ]

    def _task_exists(*extra: Any) -> Any:
        return (
            select(1)
            .select_from(ProcessingTask)
            .where(ProcessingTask.incoming_id == IncomingDocument.id, *extra)
            .exists()
        )

    open_exists = _task_exists(ProcessingTask.status != "done")
    overdue_exists = _task_exists(
        ProcessingTask.status != "done",
        ProcessingTask.deadline.is_not(None),
        ProcessingTask.deadline < today,
    )
    unprocessed = or_(open_exists, ~_task_exists())
    chua_xu_ly = (
        db.scalar(select(func.count()).select_from(IncomingDocument).where(*reg_conds, unprocessed))
        or 0
    )
    qua_han = (
        db.scalar(
            select(func.count()).select_from(IncomingDocument).where(*reg_conds, overdue_exists)
        )
        or 0
    )

    # ── Top cơ quan gửi (CV đến đã vào sổ) — luôn gồm cả 2 đơn vị (sổ đến dùng chung) ──
    sender_rows = db.execute(
        select(Organization.full_name, func.count().label("c"))
        .select_from(IncomingDocument)
        .outerjoin(Organization, Organization.id == IncomingDocument.sender_org_id)
        .where(*reg_conds)
        .group_by(Organization.full_name)
        .order_by(func.count().desc())
        .limit(7)
    ).all()
    top_senders = [{"name": r[0] or _NO_ORG, "count": int(r[1])} for r in sender_rows]

    # ── Cơ cấu loại văn bản (CV đi đã cấp số) — pie (top 7 để khớp bảng màu lát) ──
    type_rows = db.execute(
        select(DocumentType.name, func.count().label("c"))
        .select_from(OutgoingDocument)
        .join(DocumentType, DocumentType.id == OutgoingDocument.doc_type_id)
        .where(*di_conds, OutgoingDocument.issue_date >= lo, OutgoingDocument.issue_date < hi)
        .group_by(DocumentType.name)
        .order_by(func.count().desc())
        .limit(7)
    ).all()
    by_type = [{"name": r[0] or _NO_TYPE, "count": int(r[1])} for r in type_rows]

    return {
        "year": year,
        "kpi": {
            "di_year": sum(di_by_month.values()),
            "den_year": sum(den_by_month.values()),
            "di_month": di_by_month.get(cur_m, 0),
            "den_month": den_by_month.get(cur_m, 0),
            "chua_xu_ly": int(chua_xu_ly),
            "qua_han": int(qua_han),
        },
        "months": months,
        "top_senders": top_senders,
        "by_type": by_type,
    }


# --------------------------------------------------------------------------- #
# G3 — Báo cáo thống kê tuỳ chỉnh (RPT.STA)
# Lọc thời gian + đơn vị + loại VB + nhóm-theo → Excel 3 sheet (Tổng quan pivot
# + Chi tiết + Tham số). Gộp CV đi (đã cấp số) + CV đến (đã vào sổ).
# --------------------------------------------------------------------------- #

UNIT_FILTERS = ("all", "gdnn", "dvdl")
GROUP_BYS = ("month", "quarter", "sender", "type")

# Nhãn trạng thái tiếng Việt (cả 2 máy trạng thái đi/đến).
_STATUS_VN = {
    "draft": "Nháp",
    "numbered": "Đã cấp số",
    "published": "Đã phát hành",
    "registered": "Đã vào sổ",
    "cancelled": "Đã huỷ",
}
_UNIT_VN = {"all": "Tất cả", "gdnn": "Trung tâm GDNN", "dvdl": "Công ty CP DVDL"}
_GROUP_VN = {"month": "Tháng", "quarter": "Quý", "sender": "Cơ quan", "type": "Loại văn bản"}
_UNIT_CODE = {"gdnn": "GDNN", "dvdl": "DVDL"}

_NO_TYPE = "(Không phân loại)"
_NO_ORG = "(Chưa rõ cơ quan)"
_NO_PERIOD = "(Không rõ)"

# Trần số dòng chống dựng workbook quá lớn trong RAM (≈800 CV/năm → vài năm vẫn < trần).
_MAX_ROWS = 20000
_VN_OFFSET = timedelta(hours=7)  # Asia/Saigon UTC+7


def _vn_date(dt: datetime | None) -> date | None:
    """Ngày theo lịch VN từ timestamp UTC (created_at lưu UTC, trình bày/nhóm theo VN)."""
    return (dt + _VN_OFFSET).date() if dt else None


def _gather_custom_rows(
    db: Session, *, date_from: date, date_to: date, unit: str, doc_type: str
) -> list[dict[str, Any]]:
    """Nạp danh sách CV đi + đến đã chuẩn hoá cho báo cáo tuỳ chỉnh.

    Đơn vị lọc CV đi (CV đến dùng chung 2 đơn vị nên luôn gồm). Loại VB lọc theo `code`.
    `gdate` = ngày để nhóm theo tháng/quý (đi: ngày VB; đến: ngày tiếp nhận = created_at).
    """
    unit_code = _UNIT_CODE.get(unit)
    # Cận lọc CV đến theo NGÀY VN: created_at (UTC) trong [date_from .. date_to] giờ VN
    # ⇔ created_at ∈ [date_from 00:00 +07 .. (date_to+1) 00:00 +07) tính bằng UTC.
    inc_lo = datetime.combine(date_from, time.min, tzinfo=UTC) - _VN_OFFSET
    inc_hi = datetime.combine(date_to + timedelta(days=1), time.min, tzinfo=UTC) - _VN_OFFSET

    # --- CV đi ---
    out_q = (
        select(OutgoingDocument, DocumentType.code, DocumentType.name)
        .join(DocumentType, OutgoingDocument.doc_type_id == DocumentType.id)
        .where(
            OutgoingDocument.deleted_at.is_(None),
            OutgoingDocument.number_int.is_not(None),
            OutgoingDocument.issue_date >= date_from,
            OutgoingDocument.issue_date <= date_to,
        )
        .order_by(OutgoingDocument.issue_date.asc(), OutgoingDocument.number_int.asc())
    )
    if unit_code:
        uid = db.scalar(select(Unit.id).where(Unit.code == unit_code))
        out_q = out_q.where(OutgoingDocument.unit_id == (uid or -1))
    if doc_type != "all":
        out_q = out_q.where(DocumentType.code == doc_type)
    out_pairs = db.execute(out_q).all()
    out_docs = [d for d, _c, _n in out_pairs]

    # Batch người ký + nơi nhận (tránh N+1).
    profile_ids = {d.signing_profile_id for d in out_docs if d.signing_profile_id is not None}
    signer_map: dict[int, str] = {}
    if profile_ids:
        signer_map = {
            pid: name
            for pid, name in db.execute(
                select(SigningProfile.id, Signature.full_name)
                .join(Signature, SigningProfile.signature_id == Signature.id)
                .where(SigningProfile.id.in_(profile_ids))
            ).all()
        }
    recip_map: dict[int, list[str]] = {}
    if out_docs:
        for oid, name in db.execute(
            select(OutgoingRecipient.outgoing_id, Organization.full_name)
            .join(Organization, OutgoingRecipient.organization_id == Organization.id)
            .where(OutgoingRecipient.outgoing_id.in_([d.id for d in out_docs]))
        ).all():
            recip_map.setdefault(oid, []).append(name)

    rows: list[dict[str, Any]] = []
    for d, code, name in out_pairs:
        recipients = recip_map.get(d.id, [])
        rows.append({
            "direction": "Đi",
            "number": d.number or "",
            "doc_date": d.issue_date,
            "gdate": d.issue_date,
            "subject": d.subject or "",
            "type_code": code or _NO_TYPE,
            "type_name": name or _NO_TYPE,
            "signer": signer_map.get(d.signing_profile_id, "") if d.signing_profile_id else "",
            "counterpart": "; ".join(recipients),
            # Nhóm theo cơ quan: CV đi tính theo nơi nhận chính (đầu danh sách).
            "org_group": recipients[0] if recipients else "(Không có nơi nhận)",
            "status": _STATUS_VN.get(d.status, d.status),
        })

    # --- CV đến (luôn gồm; dùng chung 2 đơn vị) ---
    inc_q = (
        select(IncomingDocument, DocumentType.code, DocumentType.name)
        .outerjoin(DocumentType, IncomingDocument.doc_type_id == DocumentType.id)
        .where(
            IncomingDocument.deleted_at.is_(None),
            IncomingDocument.number_int.is_not(None),
            IncomingDocument.created_at >= inc_lo,
            IncomingDocument.created_at < inc_hi,
        )
        .order_by(IncomingDocument.created_at.asc(), IncomingDocument.number_int.asc())
    )
    if doc_type != "all":
        inc_q = inc_q.where(DocumentType.code == doc_type)
    inc_pairs = db.execute(inc_q).all()
    inc_docs = [d for d, _c, _n in inc_pairs]

    sender_ids = {d.sender_org_id for d in inc_docs if d.sender_org_id is not None}
    sender_map: dict[int, str] = {}
    if sender_ids:
        sender_map = {
            oid: nm
            for oid, nm in db.execute(
                select(Organization.id, Organization.full_name).where(
                    Organization.id.in_(sender_ids)
                )
            ).all()
        }
    for d, code, name in inc_pairs:
        sender = sender_map.get(d.sender_org_id, "") if d.sender_org_id else ""
        rows.append({
            "direction": "Đến",
            "number": d.reference_number or "",
            "doc_date": d.document_date,
            "gdate": _vn_date(d.created_at),
            "subject": d.subject or "",
            "type_code": code or _NO_TYPE,
            "type_name": name or _NO_TYPE,
            "signer": "",
            "counterpart": sender,
            "org_group": sender or _NO_ORG,
            "status": _STATUS_VN.get(d.status, d.status),
        })
    return rows


def _group_key(row: dict[str, Any], group_by: str) -> tuple[Any, str]:
    """Trả (khoá-sắp-xếp, nhãn hiển thị) cho 1 dòng theo tiêu chí nhóm."""
    if group_by == "type":
        return (row["type_code"], row["type_code"])
    if group_by == "sender":
        g = row["org_group"]
        return (g, g)
    d: date | None = row["gdate"]
    if d is None:
        return ((9999, 99), _NO_PERIOD)
    if group_by == "quarter":
        q = (d.month - 1) // 3 + 1
        return ((d.year, q), f"Quý {q}/{d.year}")
    return ((d.year, d.month), f"{d.month:02d}/{d.year}")  # month


def _render_custom_xlsx(rows: list[dict[str, Any]], *, params: dict[str, Any]) -> bytes:
    """Render Excel 3 sheet (Tổng quan pivot + Chi tiết + Tham số). Thuần — test standalone."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    group_by: str = params["group_by"]
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="FCE8B2")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- Tính pivot: cột = Loại VB (hoặc Hướng khi nhóm theo loại) ---
    if group_by == "type":
        col_keys = ["Đi", "Đến"]
        col_field = "direction"
    else:
        col_keys = sorted({r["type_code"] for r in rows})
        col_field = "type_code"

    row_order: dict[Any, str] = {}
    matrix: dict[tuple[Any, str], int] = {}
    for r in rows:
        sort_key, label = _group_key(r, group_by)
        row_order.setdefault(sort_key, label)
        ck = r[col_field]
        matrix[(sort_key, ck)] = matrix.get((sort_key, ck), 0) + 1
    sorted_rows = sorted(row_order.keys())

    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Tổng quan"

    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(col_keys) + 2)
    t = ws1.cell(row=1, column=1, value="BÁO CÁO THỐNG KÊ CÔNG VĂN")
    t.font = Font(bold=True, size=13)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 24

    hr = 3
    head = [_GROUP_VN[group_by], *[_excel_safe(c) for c in col_keys], "Tổng"]
    for c, label in enumerate(head, start=1):
        cell = ws1.cell(row=hr, column=c, value=label)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws1.column_dimensions["A"].width = 22
    for c in range(2, len(col_keys) + 3):
        ws1.column_dimensions[get_column_letter(c)].width = 14

    col_tot = {ck: 0 for ck in col_keys}
    grand = 0
    for i, sk in enumerate(sorted_rows):
        r_idx = hr + 1 + i
        ws1.cell(row=r_idx, column=1, value=_excel_safe(row_order[sk])).border = border
        rtot = 0
        for c, ck in enumerate(col_keys, start=2):
            v = matrix.get((sk, ck), 0)
            ws1.cell(row=r_idx, column=c, value=v).border = border
            rtot += v
            col_tot[ck] += v
        ws1.cell(row=r_idx, column=len(col_keys) + 2, value=rtot).border = border
        grand += rtot
    # Dòng tổng cộng.
    tr = hr + 1 + len(sorted_rows)
    tc = ws1.cell(row=tr, column=1, value="Tổng cộng")
    tc.font = bold
    tc.border = border
    for c, ck in enumerate(col_keys, start=2):
        cell = ws1.cell(row=tr, column=c, value=col_tot[ck])
        cell.font = bold
        cell.border = border
    gcell = ws1.cell(row=tr, column=len(col_keys) + 2, value=grand)
    gcell.font = bold
    gcell.border = border
    ws1.freeze_panes = ws1.cell(row=hr + 1, column=1)

    # Biểu đồ cột (tổng mỗi nhóm) — chỉ khi có dữ liệu.
    if sorted_rows:
        chart = BarChart()
        chart.title = "Số công văn theo " + _GROUP_VN[group_by].lower()
        chart.y_axis.title = "Số lượng"
        chart.x_axis.title = _GROUP_VN[group_by]
        data = Reference(ws1, min_col=len(col_keys) + 2, min_row=hr, max_row=tr - 1)
        cats = Reference(ws1, min_col=1, min_row=hr + 1, max_row=tr - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 16
        ws1.add_chart(chart, f"A{tr + 2}")

    # --- Sheet 2 Chi tiết ---
    ws2 = wb.create_sheet("Chi tiết")
    det_cols = [
        ("direction", "Hướng", 8),
        ("number", "Số, ký hiệu", 20),
        ("doc_date", "Ngày VB", 14),
        ("type_name", "Loại VB", 18),
        ("subject", "Trích yếu", 48),
        ("signer", "Người ký", 20),
        ("counterpart", "Nơi nhận / Cơ quan gửi", 32),
        ("status", "Trạng thái", 16),
    ]
    for c, (_k, label, w) in enumerate(det_cols, start=1):
        cell = ws2.cell(row=1, column=c, value=label)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.row_dimensions[1].height = 28
    for i, r in enumerate(rows, start=2):
        for c, (k, _label, _w) in enumerate(det_cols, start=1):
            val = _fmt_d(r["doc_date"]) if k == "doc_date" else r.get(k, "")
            cell = ws2.cell(row=i, column=c, value=_excel_safe(val))
            cell.border = border
            cell.alignment = Alignment(
                vertical="top", wrap_text=k in ("subject", "counterpart")
            )
    ws2.freeze_panes = ws2.cell(row=2, column=1)

    # --- Sheet 3 Tham số ---
    ws3 = wb.create_sheet("Tham số")
    ws3.column_dimensions["A"].width = 24
    ws3.column_dimensions["B"].width = 48
    meta = [
        ("Từ ngày", _fmt_d(params["date_from"])),
        ("Đến ngày", _fmt_d(params["date_to"])),
        ("Đơn vị", _UNIT_VN.get(params["unit"], params["unit"])),
        (
            "Loại văn bản",
            params.get("doc_type_label")
            or ("Tất cả" if params["doc_type"] == "all" else params["doc_type"]),
        ),
        ("Nhóm theo", _GROUP_VN.get(group_by, group_by)),
        ("Tổng số công văn", str(len(rows))),
        ("Ngày xuất báo cáo", _fmt_d(params["today"])),
    ]
    note = ws3.cell(row=1, column=1, value="BỘ LỌC ĐÃ CHỌN KHI XUẤT")
    note.font = bold
    for i, (k, v) in enumerate(meta, start=3):
        kc = ws3.cell(row=i, column=1, value=k)
        kc.font = bold
        ws3.cell(row=i, column=2, value=_excel_safe(v))
    ws3.cell(
        row=len(meta) + 4,
        column=1,
        value="Ghi chú: CV đến dùng chung 2 đơn vị nên luôn được tính, không phụ thuộc bộ lọc đơn vị.",
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_custom_report_xlsx(
    db: Session,
    *,
    date_from: date,
    date_to: date,
    unit: str,
    doc_type: str,
    group_by: str,
    today: date,
) -> bytes:
    """G3 — sinh Excel báo cáo thống kê tuỳ chỉnh (3 sheet). Trả bytes .xlsx."""
    if unit not in UNIT_FILTERS:
        raise ValidationFailed("Đơn vị không hợp lệ")
    if group_by not in GROUP_BYS:
        raise ValidationFailed("Tiêu chí nhóm không hợp lệ")
    if date_from > date_to:
        raise ValidationFailed("Khoảng thời gian không hợp lệ (Từ ngày sau Đến ngày)")
    rows = _gather_custom_rows(
        db, date_from=date_from, date_to=date_to, unit=unit, doc_type=doc_type
    )
    if len(rows) > _MAX_ROWS:
        raise ValidationFailed(
            f"Khoảng thời gian quá rộng ({len(rows):,} công văn). "
            "Vui lòng thu hẹp khoảng ngày rồi xuất lại."
        )
    # Sheet Tham số ghi TÊN loại (khớp dropdown), không phải mã.
    if doc_type == "all":
        doc_type_label = "Tất cả"
    else:
        doc_type_label = next(
            (r["type_name"] for r in rows if r["type_code"] == doc_type), doc_type
        )
    return _render_custom_xlsx(
        rows,
        params={
            "date_from": date_from,
            "date_to": date_to,
            "unit": unit,
            "doc_type": doc_type,
            "doc_type_label": doc_type_label,
            "group_by": group_by,
            "today": today,
        },
    )
