"""Unit test G2 — cấu trúc Excel sổ NĐ 30 (mở lại bằng openpyxl) + stats rỗng."""

from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any

import pytest

from app.core.errors import ValidationFailed
from app.services import report

pytestmark = pytest.mark.unit


class _Empty:
    def all(self) -> list[Any]:
        return []


class FakeDB:
    """DB rỗng — sổ không có dòng dữ liệu (kiểm header + tiêu đề chuẩn NĐ 30)."""

    def scalar(self, _stmt: Any) -> Any:
        return None

    def scalars(self, _stmt: Any) -> _Empty:
        return _Empty()

    def execute(self, _stmt: Any) -> _Empty:
        return _Empty()


def _open(data: bytes) -> Any:
    from openpyxl import load_workbook

    return load_workbook(BytesIO(data)).active


@pytest.mark.parametrize("book,cols,title_kw", [
    ("di_gdnn", report._COLS_DI, "ĐI"),
    ("di_dvdl", report._COLS_DI, "ĐI"),
    ("den", report._COLS_DEN, "ĐẾN"),
])
def test_register_xlsx_headers(book: str, cols: list[Any], title_kw: str) -> None:
    ws = _open(report.build_register_xlsx(FakeDB(), year=2026, book=book))  # type: ignore[arg-type]
    assert title_kw in ws.cell(row=1, column=1).value
    assert "2026" in ws.cell(row=1, column=1).value
    headers = [ws.cell(row=3, column=c).value for c in range(1, len(cols) + 1)]
    assert headers == [label for _key, label, _w in cols]


def test_register_xlsx_invalid_book() -> None:
    with pytest.raises(ValidationFailed):
        report.build_register_xlsx(FakeDB(), year=2026, book="xyz")  # type: ignore[arg-type]


@pytest.mark.parametrize("raw,expected", [
    ("=SUM(A1)", "'=SUM(A1)"),
    ("+1+1", "'+1+1"),
    ("-cmd", "'-cmd"),
    ("@x", "'@x"),
    ("\tTab", "'\tTab"),
    ("Bình thường", "Bình thường"),
    (5, 5),
])
def test_excel_safe(raw: object, expected: object) -> None:
    assert report._excel_safe(raw) == expected


def test_outgoing_list_xlsx_headers() -> None:
    ws = _open(report.build_outgoing_list_xlsx(FakeDB()))  # type: ignore[arg-type]
    assert "CÔNG VĂN ĐI" in ws.cell(row=1, column=1).value
    headers = [ws.cell(row=3, column=c).value for c in range(1, len(report._COLS_OUT_LIST) + 1)]
    assert headers == [label for _k, label, _w in report._COLS_OUT_LIST]


def test_incoming_list_xlsx_headers() -> None:
    ws = _open(report.build_incoming_list_xlsx(FakeDB(), include_manager_only=True))  # type: ignore[arg-type]
    assert "CÔNG VĂN ĐẾN" in ws.cell(row=1, column=1).value
    headers = [ws.cell(row=3, column=c).value for c in range(1, len(report._COLS_IN_LIST) + 1)]
    assert headers == [label for _k, label, _w in report._COLS_IN_LIST]


def test_dashboard_stats_empty() -> None:
    s = report.dashboard_stats(FakeDB(), year=2026, today=date(2026, 6, 28))  # type: ignore[arg-type]
    assert s["year"] == 2026
    assert s["kpi"] == {
        "di_year": 0, "den_year": 0, "di_month": 0, "den_month": 0,
        "chua_xu_ly": 0, "qua_han": 0,
    }
    assert len(s["months"]) == 12
    assert s["months"][0] == {"month": 1, "di": 0, "den": 0}
    assert s["top_senders"] == [] and s["by_type"] == []


# --------------------------------------------------------------------------- #
# G3 — báo cáo thống kê tuỳ chỉnh (render thuần, không cần DB)
# --------------------------------------------------------------------------- #


def _row(**kw: Any) -> dict[str, Any]:
    base: dict[str, Any] = {
        "direction": "Đi",
        "number": "001/CV",
        "doc_date": date(2026, 1, 15),
        "gdate": date(2026, 1, 15),
        "subject": "Trích yếu",
        "type_code": "CV",
        "type_name": "Công văn",
        "signer": "Nguyễn Văn A",
        "counterpart": "Sở Nội vụ",
        "org_group": "Sở Nội vụ",
        "status": "Đã phát hành",
    }
    base.update(kw)
    return base


def _params(**kw: Any) -> dict[str, Any]:
    base: dict[str, Any] = {
        "date_from": date(2026, 1, 1),
        "date_to": date(2026, 6, 30),
        "unit": "all",
        "doc_type": "all",
        "group_by": "month",
        "today": date(2026, 6, 28),
    }
    base.update(kw)
    return base


def _wb(data: bytes) -> Any:
    from openpyxl import load_workbook

    return load_workbook(BytesIO(data))


def test_custom_report_three_sheets() -> None:
    data = report._render_custom_xlsx([_row()], params=_params())
    wb = _wb(data)
    assert wb.sheetnames == ["Tổng quan", "Chi tiết", "Tham số"]


def test_custom_report_pivot_month() -> None:
    rows = [
        _row(gdate=date(2026, 1, 5), type_code="CV"),
        _row(gdate=date(2026, 1, 20), type_code="CV"),
        _row(gdate=date(2026, 2, 3), type_code="QĐ"),
    ]
    ws = _wb(report._render_custom_xlsx(rows, params=_params())).worksheets[0]
    # Header: Tháng | CV | QĐ | Tổng
    headers = [ws.cell(row=3, column=c).value for c in range(1, 5)]
    assert headers == ["Tháng", "CV", "QĐ", "Tổng"]
    # Dòng T01: 2 CV, 0 QĐ, tổng 2
    assert ws.cell(row=4, column=1).value == "01/2026"
    assert ws.cell(row=4, column=2).value == 2
    assert ws.cell(row=4, column=4).value == 2
    # Tổng cộng = 3
    assert ws.cell(row=6, column=1).value == "Tổng cộng"
    assert ws.cell(row=6, column=4).value == 3


def test_custom_report_group_by_type_pivots_direction() -> None:
    rows = [_row(direction="Đi", type_code="CV"), _row(direction="Đến", type_code="CV")]
    ws = _wb(report._render_custom_xlsx(rows, params=_params(group_by="type"))).worksheets[0]
    headers = [ws.cell(row=3, column=c).value for c in range(1, 4)]
    assert headers == ["Loại văn bản", "Đi", "Đến"]
    assert ws.cell(row=4, column=1).value == "CV"
    assert ws.cell(row=4, column=2).value == 1  # 1 Đi
    assert ws.cell(row=4, column=3).value == 1  # 1 Đến


def test_custom_report_detail_and_params_sheet() -> None:
    rows = [_row(subject="Họp giao ban", status="Đã huỷ")]
    wb = _wb(report._render_custom_xlsx(rows, params=_params(group_by="quarter")))
    detail = wb["Chi tiết"]
    assert detail.cell(row=2, column=5).value == "Họp giao ban"
    assert detail.cell(row=2, column=8).value == "Đã huỷ"
    params = wb["Tham số"]
    flat = [c.value for row in params.iter_rows() for c in row if c.value]
    assert "Nhóm theo" in flat and "Quý" in flat
    assert "01/01/2026" in flat  # Từ ngày định dạng VN


def test_custom_report_excel_safe_in_detail() -> None:
    rows = [_row(subject="=SUM(A1)", counterpart="+evil")]
    ws = _wb(report._render_custom_xlsx(rows, params=_params()))["Chi tiết"]
    assert ws.cell(row=2, column=5).value == "'=SUM(A1)"
    assert ws.cell(row=2, column=7).value == "'+evil"


@pytest.mark.parametrize("bad", [
    {"unit": "xxx"},
    {"group_by": "weekly"},
    {"date_from": date(2026, 7, 1), "date_to": date(2026, 1, 1)},
])
def test_custom_report_validation(bad: dict[str, Any]) -> None:
    kw = dict(date_from=date(2026, 1, 1), date_to=date(2026, 6, 30),
              unit="all", doc_type="all", group_by="month", today=date(2026, 6, 28))
    kw.update(bad)
    with pytest.raises(ValidationFailed):
        report.build_custom_report_xlsx(FakeDB(), **kw)  # type: ignore[arg-type]


def test_custom_report_group_by_sender_safe_label() -> None:
    # Nhóm theo cơ quan: nhãn dòng = tên cơ quan ngoài → phải qua _excel_safe.
    rows = [_row(direction="Đến", org_group="=evil()", type_code="CV")]
    ws = _wb(report._render_custom_xlsx(rows, params=_params(group_by="sender")))["Tổng quan"]
    assert ws.cell(row=3, column=1).value == "Cơ quan"
    assert ws.cell(row=4, column=1).value == "'=evil()"  # injection bị vô hiệu


def test_custom_report_gdate_none_bucketed() -> None:
    rows = [_row(gdate=None)]
    ws = _wb(report._render_custom_xlsx(rows, params=_params(group_by="month")))["Tổng quan"]
    assert ws.cell(row=4, column=1).value == "(Không rõ)"


def test_custom_report_params_uses_type_label(monkeypatch: pytest.MonkeyPatch) -> None:
    # build_custom_report_xlsx ghi TÊN loại (doc_type_label) vào sheet Tham số.
    def fake_gather(*_a: Any, **_k: Any) -> list[dict[str, Any]]:
        return [_row(type_code="QĐ", type_name="Quyết định")]

    monkeypatch.setattr(report, "_gather_custom_rows", fake_gather)
    data = report.build_custom_report_xlsx(
        FakeDB(),  # type: ignore[arg-type]
        date_from=date(2026, 1, 1), date_to=date(2026, 6, 30),
        unit="all", doc_type="QĐ", group_by="month", today=date(2026, 6, 28),
    )
    params = _wb(data)["Tham số"]
    flat = [c.value for r in params.iter_rows() for c in r if c.value]
    assert "Quyết định" in flat  # tên, không phải mã "QĐ"


def test_custom_report_empty_db_ok() -> None:
    # FakeDB rỗng → vẫn ra 3 sheet, không lỗi (pivot trống, không biểu đồ).
    data = report.build_custom_report_xlsx(
        FakeDB(),  # type: ignore[arg-type]
        date_from=date(2026, 1, 1), date_to=date(2026, 6, 30),
        unit="all", doc_type="all", group_by="month", today=date(2026, 6, 28),
    )
    assert _wb(data).sheetnames == ["Tổng quan", "Chi tiết", "Tham số"]
