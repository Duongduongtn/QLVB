"""Integration test G2 — sổ NĐ 30 + stats có dữ liệu thật (Postgres CI)."""

from __future__ import annotations

from datetime import UTC, date, datetime
from io import BytesIO

import pytest
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.document_type import DocumentType
from app.models.incoming_document import IncomingDocument
from app.models.organization import Organization
from app.models.outgoing_document import OutgoingDocument
from app.models.processing_task import ProcessingTask
from app.models.unit import Unit
from app.services import report

pytestmark = pytest.mark.integration


def _gdnn(db: Session) -> Unit:
    unit = db.scalar(select(Unit).where(Unit.code == "GDNN"))
    assert unit is not None, "seed thiếu đơn vị GDNN"
    return unit


def _doctype(db: Session, unit_id: int) -> DocumentType:
    dt = DocumentType(
        direction="out", unit_id=unit_id, name="Công văn", code="CV",
        number_format="{STT}/{NĂM}", reset_policy="year", zero_pad=3, is_active=True,
    )
    db.add(dt)
    db.flush()
    return dt


def test_outgoing_register_has_row(db_session: Session) -> None:
    unit = _gdnn(db_session)
    dt = _doctype(db_session, unit.id)
    db_session.add(OutgoingDocument(
        unit_id=unit.id, doc_type_id=dt.id, number="001/2026/CV", number_int=1,
        subject="Báo cáo quyết toán năm 2026", issue_date=date(2026, 6, 1), status="published",
    ))
    db_session.flush()

    data = report.build_register_xlsx(db_session, year=2026, book="di_gdnn")
    from openpyxl import load_workbook

    ws = load_workbook(BytesIO(data)).active
    # Dòng dữ liệu đầu (header ở row 3 → data từ row 4).
    assert ws.cell(row=4, column=1).value == 1  # STT
    assert ws.cell(row=4, column=2).value == "001/2026/CV"
    assert "quyết toán" in ws.cell(row=4, column=4).value


def test_dashboard_stats_counts(db_session: Session) -> None:
    unit = _gdnn(db_session)
    dt = _doctype(db_session, unit.id)
    db_session.add(OutgoingDocument(
        unit_id=unit.id, doc_type_id=dt.id, number="002/2026/CV", number_int=2,
        subject="CV tháng 6", issue_date=date(2026, 6, 15), status="published",
    ))
    db_session.add(IncomingDocument(
        number="0001", number_int=1, subject="CV đến test", status="registered",
    ))
    db_session.flush()

    s = report.dashboard_stats(db_session, year=2026, today=date(2026, 6, 28))
    assert s["kpi"]["di_year"] >= 1
    assert s["kpi"]["den_year"] >= 1
    assert s["kpi"]["di_month"] >= 1  # tháng 6


def test_dashboard_g1_metrics(db_session: Session) -> None:
    unit = _gdnn(db_session)
    dt = _doctype(db_session, unit.id)
    db_session.add(OutgoingDocument(
        unit_id=unit.id, doc_type_id=dt.id, number="003/2026/CV", number_int=3,
        subject="CV pie loại VB", issue_date=date(2026, 6, 10), status="published",
    ))
    org = Organization(full_name="Sở Lao động Thương binh Xã hội", is_sender=True)
    db_session.add(org)
    db_session.flush()
    inc = IncomingDocument(
        number="0002", number_int=2, subject="CV đến có việc xử lý",
        status="registered", sender_org_id=org.id,
    )
    db_session.add(inc)
    db_session.flush()
    # 1 việc đang mở + quá hạn (deadline trước "hôm nay").
    db_session.add(ProcessingTask(
        incoming_id=inc.id, unit_id=unit.id, status="in_progress", deadline=date(2026, 6, 1),
    ))
    db_session.flush()

    today = date(2026, 6, 28)
    s = report.dashboard_stats(db_session, year=2026, today=today)
    assert s["kpi"]["chua_xu_ly"] >= 1
    assert s["kpi"]["qua_han"] >= 1  # deadline 01/06 < 28/06
    assert any("Lao động" in t["name"] for t in s["top_senders"])
    assert any(t["name"] == "Công văn" and t["count"] >= 1 for t in s["by_type"])

    # Toggle đơn vị: CV đi lọc theo đơn vị (đơn vị khác id giả → 0 CV đi); còn CV ĐẾN + việc
    # xử lý dùng CHUNG 2 đơn vị (task bỏ nhãn đơn vị) → số chưa xử lý GIỐNG mọi view.
    s_other = report.dashboard_stats(db_session, year=2026, today=today, unit_id=10_000_000)
    assert s_other["kpi"]["chua_xu_ly"] == s["kpi"]["chua_xu_ly"]
    assert s_other["kpi"]["di_year"] == 0


def test_incoming_list_xlsx_respects_manager_only(db_session: Session) -> None:
    # Bất biến: NV (include_manager_only=False) KHÔNG xuất được CV "Chỉ Quản lý xem".
    from openpyxl import load_workbook

    mark = "TUYETMATXKZ"  # token hiếm để dò trong sheet
    db_session.add(IncomingDocument(
        number="0900", number_int=900, subject=f"CV mật {mark}",
        status="registered", manager_only=True,
    ))
    db_session.flush()

    def _subjects(*, manager: bool) -> str:
        data = report.build_incoming_list_xlsx(db_session, include_manager_only=manager)
        ws = load_workbook(BytesIO(data)).active
        return "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)

    assert mark not in _subjects(manager=False)  # NV: không thấy
    assert mark in _subjects(manager=True)  # Quản lý: thấy


def test_dashboard_unprocessed_edge_cases(db_session: Session) -> None:
    today = date(2026, 6, 28)
    base = report.dashboard_stats(db_session, year=2026, today=today)["kpi"]

    # (1) CV registered CHƯA giao task nào → tính 'chưa xử lý' (view toàn bộ), KHÔNG quá hạn.
    inc_unassigned = IncomingDocument(number="0101", number_int=101, subject="Chưa giao", status="registered")
    db_session.add(inc_unassigned)
    # (2) CV có task DONE → KHÔNG tính chưa xử lý.
    inc_done = IncomingDocument(number="0102", number_int=102, subject="Đã xong", status="registered")
    db_session.add(inc_done)
    # (3) CV có task mở nhưng KHÔNG deadline → chưa xử lý nhưng KHÔNG quá hạn.
    inc_nodl = IncomingDocument(number="0103", number_int=103, subject="Không hạn", status="registered")
    db_session.add(inc_nodl)
    # (4) CV đã HUỶ vào sổ với task mở → KHÔNG tính (loại nháp + huỷ qua status='registered').
    inc_cancel = IncomingDocument(number="0104", number_int=104, subject="Đã huỷ", status="cancelled")
    db_session.add(inc_cancel)
    unit = _gdnn(db_session)
    db_session.flush()
    db_session.add(ProcessingTask(incoming_id=inc_done.id, unit_id=unit.id, status="done", deadline=date(2026, 6, 1)))
    db_session.add(ProcessingTask(incoming_id=inc_nodl.id, unit_id=unit.id, status="new", deadline=None))
    db_session.add(ProcessingTask(incoming_id=inc_cancel.id, unit_id=unit.id, status="new", deadline=date(2026, 6, 1)))
    db_session.flush()

    k = report.dashboard_stats(db_session, year=2026, today=today)["kpi"]
    # +2 chưa xử lý: inc_unassigned (chưa giao) + inc_nodl (việc mở). inc_done & inc_cancel KHÔNG.
    assert k["chua_xu_ly"] == base["chua_xu_ly"] + 2
    # quá hạn KHÔNG tăng: done không tính, nodl không deadline, cancel bị loại khỏi base.
    assert k["qua_han"] == base["qua_han"]


def test_incoming_register_arrival_number_is_real(db_session: Session) -> None:
    # Cột "Số đến" phải là số đã cấp THẬT, không phải thứ tự dòng.
    db_session.add(IncomingDocument(
        number="0042", number_int=42, subject="CV đến số 42", status="registered",
        created_at=datetime(2026, 5, 2, tzinfo=UTC),
    ))
    db_session.flush()
    data = report.build_register_xlsx(db_session, year=2026, book="den")
    from openpyxl import load_workbook

    ws = load_workbook(BytesIO(data)).active
    col1 = [ws.cell(row=r, column=1).value for r in range(4, ws.max_row + 1)]
    assert "0042" in col1  # số đến thật xuất hiện ở cột 1


def test_custom_report_gathers_di_and_den(db_session: Session) -> None:
    """G3 — gộp CV đi + đến trong khoảng ngày, ra Chi tiết đủ 2 hướng."""
    unit = _gdnn(db_session)
    dt = _doctype(db_session, unit.id)
    db_session.add(OutgoingDocument(
        unit_id=unit.id, doc_type_id=dt.id, number="010/2026/CV", number_int=10,
        subject="CV đi tháng 3", issue_date=date(2026, 3, 10), status="published",
    ))
    db_session.add(IncomingDocument(
        number="0007", number_int=7, reference_number="55/SNV", subject="CV đến tháng 4",
        status="registered", created_at=datetime(2026, 4, 5, tzinfo=UTC),
    ))
    db_session.flush()

    data = report.build_custom_report_xlsx(
        db_session, date_from=date(2026, 1, 1), date_to=date(2026, 6, 30),
        unit="all", doc_type="all", group_by="month", today=date(2026, 6, 28),
    )
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(data))
    detail = wb["Chi tiết"]
    subjects = [detail.cell(row=r, column=5).value for r in range(2, detail.max_row + 1)]
    directions = {detail.cell(row=r, column=1).value for r in range(2, detail.max_row + 1)}
    assert "CV đi tháng 3" in subjects
    assert "CV đến tháng 4" in subjects
    assert directions == {"Đi", "Đến"}


def test_custom_report_unit_filter_excludes_other_di(db_session: Session) -> None:
    """Lọc đơn vị DVDL → không gồm CV đi GDNN; CV đến (chung) vẫn còn."""
    gdnn = _gdnn(db_session)
    dt = _doctype(db_session, gdnn.id)
    db_session.add(OutgoingDocument(
        unit_id=gdnn.id, doc_type_id=dt.id, number="011/2026/CV", number_int=11,
        subject="CV đi GDNN riêng", issue_date=date(2026, 2, 1), status="published",
    ))
    db_session.add(IncomingDocument(
        number="0008", number_int=8, subject="CV đến chung", status="registered",
        created_at=datetime(2026, 2, 2, tzinfo=UTC),
    ))
    db_session.flush()

    data = report.build_custom_report_xlsx(
        db_session, date_from=date(2026, 1, 1), date_to=date(2026, 6, 30),
        unit="dvdl", doc_type="all", group_by="month", today=date(2026, 6, 28),
    )
    from openpyxl import load_workbook

    detail = load_workbook(BytesIO(data))["Chi tiết"]
    subjects = [detail.cell(row=r, column=5).value for r in range(2, detail.max_row + 1)]
    assert "CV đi GDNN riêng" not in subjects  # lọc DVDL loại CV đi GDNN
    assert "CV đến chung" in subjects  # CV đến luôn gồm
